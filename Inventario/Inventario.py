import tkinter as tk
from tkcalendar import DateEntry
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import ttk, messagebox
from Inventario.Editar_Producto import mostrar_ventana_editar
import random
import string
from utils import centrar_ventana
from theme import (get_paleta, aplicar_estilos, crear_header, crear_boton,
                    crear_entry, crear_label, crear_card, crear_alerta, crear_treeview)

def verificar_vencimientos():
    archivo_excel = os.path.join("Inventario", "inventario_productos.xlsx")
    if not os.path.exists(archivo_excel):
        return []
    try:
        df = pd.read_excel(archivo_excel)
        if 'Fecha Vencimiento' not in df.columns:
            return []
        hoy = datetime.now()
        alertas = []
        for _, row in df.iterrows():
            try:
                fecha_str = str(row['Fecha Vencimiento'])
                fecha_ven = datetime.strptime(fecha_str, "%d/%m/%Y")
                dias = (fecha_ven - hoy).days
                if dias < 0:
                    alertas.append(("rojo", f"VENCIDO: {row['Nombre Producto']}"))
                elif dias <= 30:
                    alertas.append(("rojo", f"Vence en {dias} días: {row['Nombre Producto']}"))
                elif dias <= 60:
                    alertas.append(("amarillo", f"Vence en {dias} días: {row['Nombre Producto']}"))
                elif dias <= 90:
                    alertas.append(("verde", f"Vence en {dias} días: {row['Nombre Producto']}"))
            except (ValueError, TypeError):
                continue
        return alertas
    except Exception:
        return []

def verificar_stock_minimo():
    archivo_excel = os.path.join("Inventario", "inventario_productos.xlsx")
    if not os.path.exists(archivo_excel):
        return []
    try:
        df = pd.read_excel(archivo_excel)
        if 'Cantidad Producto' not in df.columns or 'Stock Minimo' not in df.columns:
            return []
        bajos = []
        for _, row in df.iterrows():
            try:
                cantidad = int(row['Cantidad Producto'])
                minimo = int(row['Stock Minimo'])
                if cantidad <= minimo:
                    bajos.append(("amarillo", f"Stock bajo: {row['Nombre Producto']} ({cantidad}/{minimo})"))
            except (ValueError, TypeError):
                continue
        return bajos
    except Exception:
        return []

def volver_al_menu(ventana):
    ventana.destroy()
    import VenMenu
    VenMenu.mostrar_ventana_menu()

def solo_enteros(char):
    return char.isdigit()

def solo_reales(char, texto):
    if char.isdigit() or (char == '.' and '.' not in texto):
        return True
    return False

def calcular_diferencia_meses(fecha_entrega, fecha_vencimiento):
    return relativedelta(fecha_vencimiento, fecha_entrega).months + relativedelta(fecha_vencimiento, fecha_entrega).years * 12

def actualizar_color_categoria(event=None):
    paleta = get_paleta()
    fecha_entrega = date_entry_entrega.get_date()
    fecha_vencimiento = date_entry_vencimiento.get_date()
    diferencia_meses = calcular_diferencia_meses(fecha_entrega, fecha_vencimiento)
    if diferencia_meses > 12:
        canvas_categoria.config(bg=paleta["alerta_verde"])
        label_categoria_texto.config(text="> 12 meses", fg=paleta["alerta_verde"])
    elif 6 <= diferencia_meses <= 12:
        canvas_categoria.config(bg=paleta["alerta_amarillo"])
        label_categoria_texto.config(text="6-12 meses", fg=paleta["alerta_amarillo"])
    else:
        canvas_categoria.config(bg=paleta["alerta_rojo"])
        label_categoria_texto.config(text="< 6 meses", fg=paleta["alerta_rojo"])

def ajustar_ancho_columnas(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

def aplicar_color_condicional(archivo_excel):
    wb = load_workbook(archivo_excel)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value == "Red":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif cell.value == "Yellow":
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            elif cell.value == "Green":
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    ajustar_ancho_columnas(ws)
    wb.save(archivo_excel)

def generar_codigo_unico():
    return ''.join(random.choices(string.ascii_lowercase + string.digits, k=6))

def mostrar_codigo_producto():
    if entry_nombre.get():
        codigo = generar_codigo_unico()
        entry_identificador.config(state="normal")
        entry_identificador.delete(0, tk.END)
        entry_identificador.insert(0, codigo)
        entry_identificador.config(state="readonly")

def guardar_datos():
    nombre_producto = entry_nombre.get().strip()
    cantidad_producto = entry_cantidad.get()
    precio_producto = entry_precio.get().strip()

    if not nombre_producto:
        tk.messagebox.showwarning("Advertencia", "Debe ingresar el nombre del producto.")
        return
    if not cantidad_producto:
        tk.messagebox.showwarning("Advertencia", "Debe ingresar la cantidad del producto.")
        return
    if not precio_producto:
        tk.messagebox.showwarning("Advertencia", "Debe ingresar el precio del producto.")
        return

    try:
        precio_producto = float(precio_producto)
    except ValueError:
        tk.messagebox.showerror("Error", "El precio ingresado no es un número válido.")
        return

    cantidad_producto = int(cantidad_producto)
    stock_minimo = entry_stock_minimo.get().strip()
    stock_minimo = int(stock_minimo) if stock_minimo else 0

    paleta = get_paleta()
    categoria_color = canvas_categoria.cget("bg")
    if categoria_color == paleta["alerta_verde"]:
        categoria = "Verde"
    elif categoria_color == paleta["alerta_amarillo"]:
        categoria = "Amarillo"
    elif categoria_color == paleta["alerta_rojo"]:
        categoria = "Rojo"
    else:
        categoria = "Sin definir"

    fecha_vencimiento = date_entry_vencimiento.get_date()
    codigo = entry_identificador.get()
    unidad = entry_unidad.get().strip() or "Unidad"

    datos = {
        "Nombre Producto": [nombre_producto + ": " + codigo],
        "Cantidad Producto": [cantidad_producto],
        "Precio Producto (COP)": [precio_producto],
        "Categoria": [categoria],
        "Fecha Vencimiento": [fecha_vencimiento.strftime("%d/%m/%Y")],
        "Stock Minimo": [stock_minimo],
        "Unidad": [unidad]
    }

    carpeta_inventario = "Inventario"
    if not os.path.exists(carpeta_inventario):
        os.makedirs(carpeta_inventario)

    archivo_excel = os.path.join(carpeta_inventario, "inventario_productos.xlsx")

    if os.path.exists(archivo_excel):
        df_existente = pd.read_excel(archivo_excel)
        if str(nombre_producto).lower() in df_existente['Nombre Producto'].str.lower().values:
            confirmacion = tk.messagebox.askyesno("Confirmación", f"El producto '{nombre_producto}' ya existe. ¿Deseas agregarlo de todos modos?")
            if not confirmacion:
                return
        df_nuevo = pd.DataFrame(datos)
        df_final = pd.concat([df_existente, df_nuevo], ignore_index=True)
        df_final.to_excel(archivo_excel, index=False)
    else:
        df = pd.DataFrame(datos)
        df.to_excel(archivo_excel, index=False)

    aplicar_color_condicional(archivo_excel)
    tk.messagebox.showinfo("Éxito", "Los datos se han guardado correctamente.")

def eliminar_producto():
    paleta = get_paleta()
    archivo = os.path.join("Inventario", "inventario_productos.xlsx")
    if not os.path.exists(archivo):
        messagebox.showwarning("Advertencia", "No hay inventario registrado.")
        return

    ventana_eliminar = tk.Toplevel()
    ventana_eliminar.title("Drogs+ - Eliminar Producto")
    ventana_eliminar.resizable(True, True)
    ventana_eliminar.minsize(620, 400)
    ventana_eliminar.configure(bg=paleta["bg_principal"])

    icon_path = os.path.join("images", "cruz_azul.ico")
    if os.path.exists(icon_path):
        ventana_eliminar.iconbitmap(icon_path)

    aplicar_estilos(ventana_eliminar)

    crear_header(ventana_eliminar, "Eliminar Producto")

    body = tk.Frame(ventana_eliminar, bg=paleta["bg_principal"])
    body.pack(fill="both", expand=True, padx=20, pady=10)

    search_card = crear_card(body, "Buscar Producto")
    search_card.pack(fill="x", pady=(0, 10))
    search_inner = tk.Frame(search_card, bg=paleta["bg_card"])
    search_inner.pack(fill="x", padx=15, pady=(0, 10))

    entry_buscar = crear_entry(search_inner)
    entry_buscar.pack(side="left", fill="x", expand=True, padx=(0, 10))
    entry_buscar.focus_set()

    columnas = ["Nombre Producto", "Cantidad Producto", "Precio Producto (COP)", "Categoria"]
    tree = crear_treeview(body, columnas)
    tree.pack(fill="both", expand=True, pady=(0, 10))

    def cargar_productos(texto=""):
        for item in tree.get_children():
            tree.delete(item)
        df = pd.read_excel(archivo)
        if texto:
            df = df[df['Nombre Producto'].str.contains(texto, case=False, na=False)]
        for i, (_, row) in enumerate(df.iterrows()):
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            tree.insert("", "end", values=[row.get(c, "") for c in columnas], tags=(tag,))

    def buscar():
        cargar_productos(entry_buscar.get().strip())

    def eliminar_seleccion():
        seleccion = tree.selection()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un producto para eliminar.")
            return
        if not messagebox.askyesno("Confirmar", f"¿Eliminar {len(seleccion)} producto(s)?"):
            return
        df = pd.read_excel(archivo)
        for item in seleccion:
            nombre = tree.item(item, "values")[0]
            df = df[df["Nombre Producto"] != nombre]
        df.to_excel(archivo, index=False)
        messagebox.showinfo("Éxito", "Productos eliminados correctamente.")
        cargar_productos(entry_buscar.get().strip())

    entry_buscar.bind("<Return>", lambda e: buscar())

    btn_frame = tk.Frame(body, bg=paleta["bg_principal"])
    btn_frame.pack(fill="x")
    crear_boton(btn_frame, "← Cerrar", ventana_eliminar.destroy, "Secundario", "pequeño").pack(side="left")
    crear_boton(btn_frame, "🔍 Buscar", buscar, "Primario", "pequeño").pack(side="left", padx=10)
    crear_boton(btn_frame, "🗑 Eliminar", eliminar_seleccion, "Peligro", "pequeño").pack(side="right")

    cargar_productos()
    centrar_ventana(ventana_eliminar)

def mostrar_ventana_inventario():
    alertas_vencimiento = verificar_vencimientos()
    alertas_stock = verificar_stock_minimo()

    paleta = get_paleta()
    ventana_inventario = tk.Tk()
    ventana_inventario.title("Drogs+ - Inventario")
    ventana_inventario.resizable(True, True)
    ventana_inventario.minsize(620, 600)
    ventana_inventario.configure(bg=paleta["bg_principal"])

    icon_path = os.path.join("images", "cruz_azul.ico")
    if os.path.exists(icon_path):
        ventana_inventario.iconbitmap(icon_path)

    aplicar_estilos(ventana_inventario)
    centrar_ventana(ventana_inventario)

    crear_header(ventana_inventario, "Gestión de Inventario")

    main_frame = tk.Frame(ventana_inventario, bg=paleta["bg_principal"])
    main_frame.pack(fill="both", expand=True, padx=25, pady=12)

    prod_card = crear_card(main_frame, "Producto")
    prod_card.pack(fill="x", pady=(0, 14))
    prod_inner = tk.Frame(prod_card, bg=paleta["bg_card"])
    prod_inner.pack(fill="x", padx=15, pady=(0, 16))

    for c in range(4):
        prod_inner.columnconfigure(c, weight=1)

    vcmd_entero = (ventana_inventario.register(solo_enteros), '%S')
    vcmd_real = (ventana_inventario.register(solo_reales), '%S', '%P')

    campos = [
        ("Nombre:", 0, 0), ("Identificador:", 0, 2),
        ("Cantidad:", 1, 0), ("Stock Mínimo:", 1, 2),
        ("Precio (COP):", 2, 0), ("Unidad:", 2, 2),
    ]
    for texto, fila, col in campos:
        crear_label(prod_inner, texto, "bold").grid(row=fila, column=col, sticky="e", padx=(0, 6), pady=6)

    global entry_nombre
    entry_nombre = crear_entry(prod_inner)
    entry_nombre.grid(row=0, column=1, sticky="ew", padx=(0, 15), pady=6)
    entry_nombre.bind("<KeyRelease>", lambda e: mostrar_codigo_producto())

    global entry_identificador
    entry_identificador = crear_entry(prod_inner, state="readonly")
    entry_identificador.grid(row=0, column=3, sticky="ew", pady=6)

    global entry_cantidad
    entry_cantidad = crear_entry(prod_inner, validate="key", validatecommand=vcmd_entero)
    entry_cantidad.grid(row=1, column=1, sticky="ew", padx=(0, 15), pady=6)

    global entry_stock_minimo
    entry_stock_minimo = crear_entry(prod_inner, validate="key", validatecommand=vcmd_entero)
    entry_stock_minimo.grid(row=1, column=3, sticky="ew", pady=6)

    global entry_precio
    entry_precio = crear_entry(prod_inner, validate="key", validatecommand=vcmd_real)
    entry_precio.grid(row=2, column=1, sticky="ew", padx=(0, 15), pady=6)

    global entry_unidad
    entry_unidad = crear_entry(prod_inner)
    entry_unidad.grid(row=2, column=3, sticky="ew", pady=6)
    entry_unidad.insert(0, "Unidad")

    fechas_card = crear_card(main_frame, "Fechas y Categoría")
    fechas_card.pack(fill="x", pady=(0, 14))
    fechas_inner = tk.Frame(fechas_card, bg=paleta["bg_card"])
    fechas_inner.pack(fill="x", padx=15, pady=(0, 16))

    for c in range(4):
        fechas_inner.columnconfigure(c, weight=1)

    crear_label(fechas_inner, "Fecha Entrega:", "bold").grid(row=0, column=0, sticky="e", padx=(0, 6), pady=6)
    global date_entry_entrega
    date_entry_entrega = DateEntry(fechas_inner, date_pattern='dd/mm/yyyy', width=18, background='darkblue',
                                   foreground='white', borderwidth=2, state='readonly')
    date_entry_entrega.grid(row=0, column=1, sticky="w", padx=(0, 15), pady=6)
    date_entry_entrega.bind("<<DateEntrySelected>>", actualizar_color_categoria)
    crear_label(fechas_inner, "Fecha Vencimiento:", "bold").grid(row=0, column=2, sticky="e", padx=(0, 6), pady=6)
    global date_entry_vencimiento
    date_entry_vencimiento = DateEntry(fechas_inner, date_pattern='dd/mm/yyyy', width=18, background='darkblue',
                                       foreground='white', borderwidth=2, state='readonly')
    date_entry_vencimiento.grid(row=0, column=3, sticky="w", pady=6)
    date_entry_vencimiento.bind("<<DateEntrySelected>>", actualizar_color_categoria)

    crear_label(fechas_inner, "Categoría:", "bold").grid(row=1, column=0, sticky="e", padx=(0, 6), pady=6)
    global canvas_categoria
    canvas_categoria = tk.Canvas(fechas_inner, width=100, height=22, bg=paleta["alerta_verde"],
                                  highlightthickness=1, highlightbackground=paleta["borde"])
    canvas_categoria.grid(row=1, column=1, sticky="w", pady=6)
    global label_categoria_texto
    label_categoria_texto = crear_label(fechas_inner, "", "info")
    label_categoria_texto.grid(row=1, column=2, columnspan=2, sticky="w", padx=(10, 0), pady=6)
    actualizar_color_categoria()

    if alertas_vencimiento or alertas_stock:
        alertas_card = crear_card(main_frame, "Alertas")
        alertas_card.pack(fill="x", pady=(0, 14))

        colores_alerta = {
            "rojo": (paleta["alerta_rojo"], paleta["alerta_rojo_bg"], "●"),
            "amarillo": (paleta["alerta_amarillo"], paleta["alerta_amarillo_bg"], "●"),
            "verde": (paleta["alerta_verde"], paleta["alerta_verde_bg"], "●"),
        }

        for tipo, texto in (alertas_vencimiento + alertas_stock)[:6]:
            color_texto, color_fondo, icono = colores_alerta.get(tipo, colores_alerta["rojo"])
            fila = tk.Frame(alertas_card, bg=color_fondo)
            fila.pack(fill="x", padx=10, pady=2)
            barra = tk.Frame(fila, bg=color_texto, width=4)
            barra.pack(side="left", fill="y")
            tk.Label(fila, text=f" {icono} ", font=("Segoe UI", 10), fg=color_texto,
                     bg=color_fondo).pack(side="left")
            tk.Label(fila, text=texto, font=("Segoe UI", 9), fg=color_texto,
                     bg=color_fondo, anchor="w").pack(side="left", fill="x", expand=True, padx=(0, 10), pady=6)

    btn_frame = tk.Frame(main_frame, bg=paleta["bg_principal"])
    btn_frame.pack(fill="x", pady=(8, 0))

    crear_boton(btn_frame, "← Volver", lambda: volver_al_menu(ventana_inventario), "Secundario", "pequeño").pack(side="left")
    crear_boton(btn_frame, "🔍 Buscar", mostrar_ventana_editar, "Secundario", "pequeño").pack(side="left", padx=(10, 0))
    crear_boton(btn_frame, "✏️ Editar", mostrar_ventana_editar, "Advertencia", "pequeño").pack(side="left", padx=(10, 0))
    crear_boton(btn_frame, "🗑 Eliminar", eliminar_producto, "Peligro", "pequeño").pack(side="left", padx=(10, 0))
    crear_boton(btn_frame, "💾 Guardar", guardar_datos, "Exito", "pequeño").pack(side="right")

    entry_nombre.focus_set()
    centrar_ventana(ventana_inventario)
    ventana_inventario.mainloop()
